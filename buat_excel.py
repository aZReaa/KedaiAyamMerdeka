"""
Buat file Excel siap lampiran skripsi:
- Sheet 1: Data Uji NLU (No, Pesan, Predicted, Actual, Keterangan)
- Sheet 2: Hasil Metrik (Precision, Recall, F1-Score per intent)
- Sheet 3: Ringkasan Macro Average
"""
import csv
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ======================== DATA ========================

# Urutan pesan sesuai seed_chat_logs.py
PESAN = [
    # SALAM (20)
    "halo", "hai", "pole-pole kak", "elo kak, ada ji?", "iyye assalamualaikum",
    "halo kak, engka ji disini?", "selamat pagi kak", "malam kak", "hei bot ada ko?",
    "hello kak", "siang kak", "selamat sore", "permisi kak", "hai kak engka ki?",
    "numpang tanya kak", "adekko kak!", "assalamualaikum kak", "hay kak!", "selamat datang",
    "apa khabar kak",
    # PESAN_MENU (25)
    "melo pesan ayam geprek kak", "mau ji pesan 2 ayam bakar", "order ayam goreng 1 porsi kak",
    "beli ji ayam geprek kak", "pesan ji ayam bakar sama nasi", "melo ji ayam goreng 3 kak",
    "kasih mi ayam crispy 2", "ayam geprek 1 sambal ijo pale", "mau ji order kak",
    "pesan es teh manis kak", "2 nasi sama es jeruk pale", "order mi ayam geprek pole tanpa sambal",
    "melo beli ayam apa ji ada?", "tiga ji ayam bakar kak", "pesan tempe crispy 2 kak",
    "1 ayam goreng 1 nasi pale kak", "melo makang ayam geprek", "order 2 es campur kak",
    "beli es teler 1 pale", "tambah ji 1 nasi lagi kak", "minta ayam geprek 2 sambal bawang pale",
    "pesan ji tahu crispy kak", "1 ayam bakar pale kak", "melo mapperi makang kak",
    "2 ayam geprek 1 nasi pale",
    # CEK_STATUS (10)
    "cek ji pesanan ku kak", "pura ji di proses pesanan ku?", "makanan ku pura selesai mi?",
    "pesanan ku mana ji kak?", "udah selesai mi kak?", "cek status pale kak",
    "de pura ji kak pesanan ku?", "kapan mi selesai kak?", "lama pi kak",
    "pesanan ku aga nomornya?",
    # TERIMA_KASIH (8)
    "makasih ji kak", "terima kasih kak", "thanks kak", "makasih banyak kak",
    "sip makasih kak", "oke tengkyu kak", "iyye makasih pole kak", "siap kak makasih ji",
    # INFO_PEMBAYARAN (8)
    "bayar pakai apa ji kak?", "cara bayar gimana kak?", "bisa bayar di tempat ki kak?",
    "engka ji gopay?", "terima ji transfer kak?", "pembayaran apa ji ada kak?",
    "ovo bisa ji kak?", "cod bisa kak?",
    # INFO_JAM & LOKASI (6)
    "jam berapa ji buka kak?", "tutup jam berapa kak?", "alamat ji kak dimana?",
    "lokasi ko kak?", "masih buka mi kak?", "ditempat aga kak?",
    # BATALKAN / UBAH (5)
    "batal ji pesanan ku kak", "de jadi mi kak, cancel pale", "ganti ji pesanan ku kak",
    "edit pale pesanan ku kak", "de jadi kak pesan mi",
    # UNKNOWN & REKOMENDASI (8)
    "aga ko kak?", "engka pallawa?", "cuaca sore ini kak?", "harga minyak goreng kak?",
    "mammak de engka?", "boleh minta nomormu kak?", "rekomen aga mi kak?", "menu aga ji ada kak?",
]

LABELED = [
    ("salam","salam"),("salam","salam"),("salam","salam"),("salam","salam"),("salam","salam"),
    ("salam","salam"),("salam","salam"),("salam","salam"),("salam","salam"),("salam","salam"),
    ("salam","salam"),("salam","salam"),("salam","salam"),
    ("unknown","salam"),("unknown","salam"),("unknown","salam"),
    ("salam","salam"),("salam","salam"),("salam","salam"),("unknown","salam"),

    ("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),
    ("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),
    ("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),
    ("salam","pesan_menu"),
    ("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),
    ("salam","pesan_menu"),
    ("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),
    ("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),
    ("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),("pesan_menu","pesan_menu"),
    ("pesan_menu","pesan_menu"),
    ("unknown","pesan_menu"),
    ("pesan_menu","pesan_menu"),

    ("cek_status","cek_status"),("cek_status","cek_status"),("cek_status","cek_status"),
    ("cek_status","cek_status"),("cek_status","cek_status"),("cek_status","cek_status"),
    ("unknown","cek_status"),
    ("cek_status","cek_status"),
    ("unknown","cek_status"),
    ("cek_status","cek_status"),

    ("terima_kasih","terima_kasih"),("terima_kasih","terima_kasih"),
    ("terima_kasih","terima_kasih"),("terima_kasih","terima_kasih"),
    ("terima_kasih","terima_kasih"),("terima_kasih","terima_kasih"),
    ("salam","terima_kasih"),
    ("terima_kasih","terima_kasih"),

    ("info_pembayaran","info_pembayaran"),("info_pembayaran","info_pembayaran"),
    ("info_pembayaran","info_pembayaran"),
    ("unknown","info_pembayaran"),
    ("info_pembayaran","info_pembayaran"),("info_pembayaran","info_pembayaran"),
    ("info_pembayaran","info_pembayaran"),("info_pembayaran","info_pembayaran"),

    ("info_jam","info_jam"),("info_jam","info_jam"),
    ("lokasi","lokasi"),("lokasi","lokasi"),("info_jam","info_jam"),
    ("unknown","lokasi"),

    ("batalkan_pesanan","batalkan_pesanan"),
    ("unknown","batalkan_pesanan"),
    ("ubah_pesanan","ubah_pesanan"),("ubah_pesanan","ubah_pesanan"),
    ("batalkan_pesanan","batalkan_pesanan"),

    ("unknown","unknown"),("unknown","unknown"),("unknown","unknown"),
    ("unknown","unknown"),("unknown","unknown"),("unknown","unknown"),
    ("rekomendasi_menu","rekomendasi_menu"),("rekomendasi_menu","rekomendasi_menu"),
]

# ======================== STYLE HELPERS ========================

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

# ======================== BUAT WORKBOOK ========================

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: DATA UJI NLU
# ============================================================
ws1 = wb.active
ws1.title = "Data Uji NLU"

# Judul
ws1.merge_cells("A1:F1")
ws1["A1"] = "Data Pengujian NLU Chatbot Kedai Ayam Merdeka"
ws1["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws1["A1"].fill = header_fill("1F4E79")
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 30

# Sub-judul
ws1.merge_cells("A2:F2")
ws1["A2"] = "Campuran Bahasa Indonesia dan Bahasa Luwu/Palopo (90 Sampel)"
ws1["A2"].font = Font(italic=True, size=10, color="FFFFFF")
ws1["A2"].fill = header_fill("2E75B6")
ws1["A2"].alignment = center()

# Header kolom
headers = ["No", "Pesan Masuk (Input)", "Intent Prediksi Sistem", "Intent Aktual", "Keterangan", "Kosakata Daerah"]
col_widths = [5, 42, 24, 24, 12, 22]
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=3, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = header_fill("2E75B6")
    cell.alignment = center(wrap=True)
    cell.border = thin_border()
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.row_dimensions[3].height = 28

# Deteksi kosakata daerah Luwu/Palopo
KATA_DAERAH = ["ji","mi","ko","pi","mo","to","pole","melo","elo","engka","iyye",
               "aga","de","pura","pale","mapperi","makang","adekko","pallawa","mammak"]

def deteksi_daerah(pesan):
    found = [k for k in KATA_DAERAH if k in pesan.lower().split() or f" {k}" in pesan.lower() or pesan.lower().startswith(k)]
    return ", ".join(set(found)) if found else "-"

# Isi data
FILL_BENAR = PatternFill("solid", fgColor="E2EFDA")
FILL_SALAH = PatternFill("solid", fgColor="FCE4D6")

for i, ((pred, actual), pesan) in enumerate(zip(LABELED, PESAN), 1):
    row = i + 3
    benar = pred == actual
    ket   = "Benar" if benar else "Salah"
    daerah = deteksi_daerah(pesan)
    fill  = FILL_BENAR if benar else FILL_SALAH

    vals = [i, pesan, pred.replace("_", " "), actual.replace("_", " "), ket, daerah]
    for col, val in enumerate(vals, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.fill  = fill
        cell.border = thin_border()
        cell.alignment = center() if col in [1, 5] else left()
        cell.font = Font(size=9, bold=(col==5 and not benar))

ws1.freeze_panes = "A4"

# Ringkasan kecil di kanan
summary_row = 3
ws1.cell(row=summary_row, column=8, value="RINGKASAN").font = Font(bold=True, size=10)
ws1.cell(row=4, column=8, value="Total Sampel").font = Font(size=9)
ws1.cell(row=4, column=9, value=len(LABELED)).font = Font(size=9, bold=True)
benar_count = sum(1 for p,a in LABELED if p==a)
ws1.cell(row=5, column=8, value="Prediksi Benar").font = Font(size=9, color="375623")
ws1.cell(row=5, column=9, value=benar_count).font = Font(size=9, bold=True, color="375623")
ws1.cell(row=6, column=8, value="Prediksi Salah").font = Font(size=9, color="C00000")
ws1.cell(row=6, column=9, value=len(LABELED)-benar_count).font = Font(size=9, bold=True, color="C00000")
acc = round(benar_count/len(LABELED)*100, 2)
ws1.cell(row=7, column=8, value="Akurasi").font = Font(size=9)
ws1.cell(row=7, column=9, value=f"{acc}%").font = Font(size=9, bold=True)
ws1.column_dimensions["H"].width = 18
ws1.column_dimensions["I"].width = 10

# ============================================================
# SHEET 2: HASIL METRIK PER INTENT
# ============================================================
ws2 = wb.create_sheet("Hasil Metrik")

ws2.merge_cells("A1:H1")
ws2["A1"] = "Hasil Evaluasi Teknis - Precision, Recall, F1-Score per Intent"
ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws2["A1"].fill = header_fill("1F4E79")
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:H2")
ws2["A2"] = "Metode: Fuzzy Matching NLU | Total Sampel: 90 | Campuran Bahasa Indonesia + Luwu/Palopo"
ws2["A2"].font = Font(italic=True, size=10, color="FFFFFF")
ws2["A2"].fill = header_fill("2E75B6")
ws2["A2"].alignment = center()

h2 = ["No", "Intent", "TP", "FP", "FN", "Precision", "Recall", "F1-Score"]
h2_w = [5, 22, 6, 6, 6, 12, 12, 12]
for col, (h, w) in enumerate(zip(h2, h2_w), 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = header_fill("2E75B6")
    cell.alignment = center(wrap=True)
    cell.border = thin_border()
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.row_dimensions[3].height = 28

# Hitung metrik per intent
intents = sorted(set(a for _,a in LABELED))
FILL_ROWS = [PatternFill("solid", fgColor="DEEAF1"), PatternFill("solid", fgColor="FFFFFF")]

metrik_all = []
for i, intent in enumerate(intents):
    tp = sum(1 for p,a in LABELED if p==intent and a==intent)
    fp = sum(1 for p,a in LABELED if p==intent and a!=intent)
    fn = sum(1 for p,a in LABELED if p!=intent and a==intent)
    pr = tp/(tp+fp) if (tp+fp)>0 else 0
    rc = tp/(tp+fn) if (tp+fn)>0 else 0
    f1 = 2*pr*rc/(pr+rc) if (pr+rc)>0 else 0
    metrik_all.append((intent, tp, fp, fn, pr, rc, f1))

    row = i + 4
    fill = FILL_ROWS[i % 2]
    vals = [i+1, intent.replace("_"," "), tp, fp, fn,
            f"{pr*100:.2f}%", f"{rc*100:.2f}%", f"{f1*100:.2f}%"]
    for col, val in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = thin_border()
        cell.alignment = center() if col != 2 else left(wrap=False)
        cell.font = Font(size=10)

# Macro Average
macro_row = len(intents) + 4
mp = sum(m[4] for m in metrik_all)/len(metrik_all)
mr = sum(m[5] for m in metrik_all)/len(metrik_all)
mf = sum(m[6] for m in metrik_all)/len(metrik_all)

macro_vals = ["", "MACRO AVERAGE", "", "", "",
              f"{mp*100:.2f}%", f"{mr*100:.2f}%", f"{mf*100:.2f}%"]
for col, val in enumerate(macro_vals, 1):
    cell = ws2.cell(row=macro_row, column=col, value=val)
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = header_fill("1F4E79")
    cell.border = thin_border()
    cell.alignment = center()

ws2.freeze_panes = "A4"

# Catatan
note_row = macro_row + 2
ws2.merge_cells(f"A{note_row}:H{note_row}")
ws2[f"A{note_row}"] = "Keterangan: TP=True Positive | FP=False Positive | FN=False Negative"
ws2[f"A{note_row}"].font = Font(italic=True, size=9, color="595959")
ws2[f"A{note_row}"].alignment = left()

ws2.merge_cells(f"A{note_row+1}:H{note_row+1}")
ws2[f"A{note_row+1}"] = f"Error terbanyak pada intent 'salam' dan 'unknown' akibat kosakata Luwu/Palopo (engka, de, pole, mapperi, melo) yang tidak dikenali NLU berbasis fuzzy matching Bahasa Indonesia."
ws2[f"A{note_row+1}"].font = Font(italic=True, size=9, color="595959")
ws2[f"A{note_row+1}"].alignment = left(wrap=True)
ws2.row_dimensions[note_row+1].height = 30

# ============================================================
# SHEET 3: RUMUS METRIK
# ============================================================
ws3 = wb.create_sheet("Rumus & Keterangan")

ws3.merge_cells("A1:C1")
ws3["A1"] = "Rumus Metrik Evaluasi Teknis Chatbot"
ws3["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws3["A1"].fill = header_fill("1F4E79")
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 28
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 35
ws3.column_dimensions["C"].width = 45

rows = [
    ("Metrik", "Rumus", "Keterangan"),
    ("Precision", "TP / (TP + FP)", "Ketepatan sistem dalam memprediksi suatu intent. Tinggi = jarang salah prediksi positif."),
    ("Recall", "TP / (TP + FN)", "Kelengkapan sistem mengenali intent. Tinggi = sistem jarang melewatkan intent yang benar."),
    ("F1-Score", "2 × (P × R) / (P + R)", "Rata-rata harmonik Precision dan Recall. Keseimbangan antara ketepatan dan kelengkapan."),
    ("Macro Avg", "Rata-rata seluruh intent", "Setiap intent diperlakukan setara tanpa melihat jumlah sampel masing-masing."),
]
for i, (a, b, c) in enumerate(rows):
    row = i + 2
    is_header = i == 0
    fill = header_fill("2E75B6") if is_header else FILL_ROWS[i % 2]
    font_style = Font(bold=True, color="FFFFFF", size=10) if is_header else Font(size=10)
    for col, val in enumerate([a, b, c], 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = font_style
        cell.fill = fill
        cell.border = thin_border()
        cell.alignment = center() if col < 3 else left(wrap=True)
    ws3.row_dimensions[row].height = 40

# ======================== SIMPAN ========================
out = "Lampiran_Evaluasi_Chatbot.xlsx"
wb.save(out)
import os
print(f"[OK] File Excel berhasil dibuat: {os.path.abspath(out)}")
print(f"     - Sheet 1: Data Uji NLU (90 sampel + warna benar/salah)")
print(f"     - Sheet 2: Hasil Metrik per Intent")
print(f"     - Sheet 3: Rumus & Keterangan")
